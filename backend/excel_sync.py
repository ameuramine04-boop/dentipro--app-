"""
Miroir de utils/excel.js — écriture thread-safe dans dentipro_data.xlsx.
"""
from __future__ import annotations

import threading
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_executor = None
_ROOT = Path(__file__).resolve().parent.parent
EXCEL_FILE = _ROOT / "dentipro_data.xlsx"

T = {
    "Patients": {"h": "1A73E8", "f": "FFFFFF", "tab": "1A73E8", "e": "E8F0FE", "o": "FFFFFF"},
    "Rdv": {"h": "0F9D58", "f": "FFFFFF", "tab": "0F9D58", "e": "E6F4EA", "o": "FFFFFF"},
    "Paiements": {"h": "7B1FA2", "f": "FFFFFF", "tab": "7B1FA2", "e": "F3E5F5", "o": "FFFFFF"},
    "Factures": {
        "h": "E65100",
        "f": "FFFFFF",
        "tab": "E65100",
        "e": "FFF3E0",
        "o": "FFFFFF",
        "st": {
            "Payee": {"f": "2E7D32", "b": "E8F5E9"},
            "Impayee": {"f": "C62828", "b": "FFEBEE"},
            "Partiellement payee": {"f": "F57F17", "b": "FFF9C4"},
        },
    },
    "Historique": {"h": "37474F", "f": "FFFFFF", "tab": "37474F", "e": "F5F5F5", "o": "FFFFFF"},
}

HC = {
    "PATIENT_AJOUT": {"b": "E3F2FD", "f": "1565C0", "i": "\U0001f464"},
    "PATIENT_MODIF": {"b": "FCE4EC", "f": "C62828", "i": "\u270f\ufe0f"},
    "PATIENT_SUPPRIM": {"b": "FFEBEE", "f": "B71C1C", "i": "\U0001f5d1\ufe0f"},
    "RDV_AJOUT": {"b": "E8F5E9", "f": "2E7D32", "i": "\U0001f4c5"},
    "RDV_MODIF": {"b": "FFF9C4", "f": "F57F17", "i": "\u270f\ufe0f"},
    "RDV_SUPPRIM": {"b": "FFEBEE", "f": "C62828", "i": "\U0001f5d1\ufe0f"},
    "PAIEMENT_AJOUT": {"b": "F3E5F5", "f": "6A1B9A", "i": "\U0001f4b0"},
    "FACTURE_AJOUT": {"b": "FFF3E0", "f": "E65100", "i": "\U0001f4c4"},
    "FACTURE_MAJ": {"b": "FBE9E7", "f": "D84315", "i": "\U0001f504"},
    "FACTURE_SUPPRIM": {"b": "FFEBEE", "f": "C62828", "i": "\U0001f5d1\ufe0f"},
    "SALLE_AJOUT": {"b": "E8EAF6", "f": "283593", "i": "\U0001fa91"},
    "DEFAULT": {"b": "FFFFFF", "f": "424242", "i": "\U0001f4dd"},
}

COLS = {
    "Patients": [
        {"header": "ID", "width": 7},
        {"header": "Nom", "width": 18},
        {"header": "Prénom", "width": 18},
        {"header": "Sexe", "width": 10},
        {"header": "Téléphone", "width": 16},
        {"header": "CNIE", "width": 14},
        {"header": "Date Naissance", "width": 16},
        {"header": "Email", "width": 26},
        {"header": "Créé le", "width": 20},
    ],
    "Rdv": [
        {"header": "ID", "width": 7},
        {"header": "Patient ID", "width": 10},
        {"header": "Patient", "width": 24},
        {"header": "Date", "width": 14},
        {"header": "Heure", "width": 10},
        {"header": "Motif", "width": 28},
        {"header": "Dent(s)", "width": 12},
        {"header": "Statut", "width": 14},
        {"header": "Créé le", "width": 20},
    ],
    "Paiements": [
        {"header": "ID", "width": 7},
        {"header": "Patient ID", "width": 10},
        {"header": "Patient", "width": 24},
        {"header": "Montant (DH)", "width": 14},
        {"header": "Type", "width": 12},
        {"header": "Date", "width": 14},
        {"header": "Facture N°", "width": 14},
        {"header": "Facture ID", "width": 10},
        {"header": "Notes", "width": 24},
        {"header": "Créé le", "width": 20},
    ],
    "Factures": [
        {"header": "ID", "width": 7},
        {"header": "Patient ID", "width": 10},
        {"header": "Patient", "width": 24},
        {"header": "N° Facture", "width": 14},
        {"header": "Date", "width": 14},
        {"header": "Motif", "width": 20},
        {"header": "Dent(s)", "width": 12},
        {"header": "Total DH", "width": 14},
        {"header": "Payé DH", "width": 14},
        {"header": "Reste DH", "width": 14},
        {"header": "Statut", "width": 22},
        {"header": "Dernière MAJ", "width": 20},
    ],
    "Historique": [
        {"header": "#", "width": 5},
        {"header": "Date/Heure", "width": 20},
        {"header": "Type", "width": 22},
        {"header": "Opération", "width": 28},
        {"header": "Patient", "width": 24},
        {"header": "Détails", "width": 42},
        {"header": "Utilisateur", "width": 18},
    ],
}

_lock = threading.Lock()


def _argb(hex6_or_8: str) -> str:
    h = (hex6_or_8 or "FFFFFF").upper().replace("#", "")
    if len(h) == 8:
        return h
    return "FF" + h[-6:].rjust(6, "0")


def configure(execute_fn):
    global _executor
    _executor = execute_fn


def _q(sql: str, params=None):
    if _executor is None:
        raise RuntimeError("excel_sync.configure() must be called from app")
    return _executor(sql, params or ())


def fd(v):
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, date):
        return v.strftime("%d/%m/%Y")
    s = str(v)
    try:
        if "T" in s:
            return datetime.fromisoformat(s.replace("Z", "+00:00")).strftime("%d/%m/%Y")
        if len(s) >= 10 and s[4] == "-":
            return datetime.strptime(s[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        pass
    return s


def fdt(v):
    if not v:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y %H:%M")
    if isinstance(v, date):
        return v.strftime("%d/%m/%Y")
    try:
        d = datetime.fromisoformat(str(v).replace("Z", "+00:00"))
        return d.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(v)


def _thin():
    return Side(style="thin", color="FFCCCCCC")


def _style_header(ws, name):
    t = T[name]
    fill = PatternFill("solid", fgColor=_argb(t["h"]))
    font = Font(bold=True, color=_argb(t["f"]), size=11, name="Calibri")
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = Border(top=_thin(), left=_thin(), bottom=_thin(), right=_thin())
    ws.row_dimensions[1].height = 24


def _style_data_row(ws, rn, name, fill_hex=None):
    t = T[name]
    raw = fill_hex or (t["e"] if rn % 2 == 0 else t["o"])
    pf = PatternFill("solid", fgColor=_argb(raw))
    font = Font(size=10, name="Calibri")
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=rn, column=c)
        cell.fill = pf
        cell.font = font
        cell.alignment = Alignment(vertical="center")
        cell.border = Border(top=_thin(), left=_thin(), bottom=_thin(), right=_thin())
    ws.row_dimensions[rn].height = 18


def _get_wb():
    from openpyxl.utils import get_column_letter

    if EXCEL_FILE.exists():
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()
        wb.remove(wb.active)
    wb.properties.creator = "DentiPro"
    for name, cols in COLS.items():
        if name not in wb.sheetnames:
            ws = wb.create_sheet(title=name)
            for i, col in enumerate(cols, 1):
                ws.cell(row=1, column=i, value=col["header"])
                ws.column_dimensions[get_column_letter(i)].width = col["width"]
            _style_header(ws, name)
    return wb


def _hist(wb, typ, op, patient, details, user="Système"):
    ws = wb["Historique"]
    rn = ws.max_row + 1
    c = HC.get(typ, HC["DEFAULT"])
    label = typ.replace("_", " ")
    ws.cell(row=rn, column=1, value=rn - 1)
    ws.cell(row=rn, column=2, value=fdt(datetime.now()))
    ws.cell(row=rn, column=3, value=f"{c['i']} {label}")
    ws.cell(row=rn, column=4, value=op)
    ws.cell(row=rn, column=5, value=patient or "")
    ws.cell(row=rn, column=6, value=details or "")
    ws.cell(row=rn, column=7, value=user or "Système")
    _style_data_row(ws, rn, "Historique", c["b"])
    ws.cell(row=rn, column=3).font = Font(bold=True, color=_argb(c["f"]), size=10, name="Calibri")
    ws.cell(row=rn, column=4).font = Font(bold=True, color=_argb(c["f"]), size=10, name="Calibri")


def _with_lock_sync(fn):
    with _lock:
        fn()


def append_patient(patient_id: int):
    rows = _q("SELECT * FROM patient WHERE id_patient=%s", (patient_id,))
    if not rows:
        return
    p = rows[0]

    def job():
        wb = _get_wb()
        ws = wb["Patients"]
        rn = ws.max_row + 1
        ws.append(
            [
                p.get("id_patient"),
                p.get("nom"),
                p.get("prenom"),
                p.get("sexe") or "",
                p.get("telephone") or "",
                p.get("cnie") or "",
                fd(p.get("date_naissance")),
                p.get("email") or "",
                fdt(p.get("created_at")),
            ]
        )
        _style_data_row(ws, rn, "Patients")
        _hist(
            wb,
            "PATIENT_AJOUT",
            "Nouveau patient",
            f"{p.get('nom')} {p.get('prenom')}",
            f"Tél: {p.get('telephone') or '-'} | CIN: {p.get('cnie') or '-'}",
        )
        wb.save(EXCEL_FILE)

    _with_lock_sync(job)


def append_rdv(rdv_id: int):
    rows = _q(
        "SELECT r.*, CONCAT(p.nom,' ',p.prenom) AS pn FROM rendez_vous r "
        "LEFT JOIN patient p ON r.id_patient=p.id_patient WHERE r.id_rdv=%s",
        (rdv_id,),
    )
    if not rows:
        return
    r = rows[0]

    def job():
        wb = _get_wb()
        ws = wb["Rdv"]
        rn = ws.max_row + 1
        ws.append(
            [
                r.get("id_rdv"),
                r.get("id_patient"),
                r.get("pn"),
                fd(r.get("date_rdv")),
                str(r.get("heure_rdv") or ""),
                r.get("motif") or "",
                r.get("dent") or "",
                r.get("statut") or "",
                fdt(r.get("created_at")),
            ]
        )
        sf = {"Prevu": "FFF9C4", "En cours": "E3F2FD", "Termine": "E8F5E9", "Annule": "FFEBEE"}
        st = r.get("statut") or ""
        _style_data_row(ws, rn, "Rdv", sf.get(st) or None)
        _hist(
            wb,
            "RDV_AJOUT",
            "RDV planifié",
            r.get("pn"),
            f"{fd(r.get('date_rdv'))} à {r.get('heure_rdv') or '?'} — {r.get('motif') or 'Consultation'}",
        )
        wb.save(EXCEL_FILE)

    _with_lock_sync(job)


def append_paiement(pay_id: int):
    rows = _q(
        "SELECT p.*, CONCAT(pt.nom,' ',pt.prenom) AS pn, f.numero_facture FROM paiement p "
        "LEFT JOIN patient pt ON p.id_patient=pt.id_patient "
        "LEFT JOIN facture f ON p.id_facture=f.id_facture WHERE p.id_paiement=%s",
        (pay_id,),
    )
    if not rows:
        return
    p = rows[0]

    def job():
        wb = _get_wb()
        ws = wb["Paiements"]
        rn = ws.max_row + 1
        m = float(p.get("montant") or 0)
        ws.append(
            [
                p.get("id_paiement"),
                p.get("id_patient"),
                p.get("pn"),
                m,
                p.get("type_paiement") or "",
                fd(p.get("date_paiement")),
                p.get("numero_facture") or "",
                p.get("id_facture") or "",
                p.get("notes") or "",
                fdt(p.get("created_at")),
            ]
        )
        _style_data_row(ws, rn, "Paiements")
        ws.cell(row=rn, column=4).number_format = '#,##0.00 "DH"'
        fac = p.get("numero_facture")
        _hist(
            wb,
            "PAIEMENT_AJOUT",
            "Paiement reçu",
            p.get("pn"),
            f"{m:.2f} DH — {p.get('type_paiement') or 'Especes'}"
            + (f" | Fac.{fac}" if fac else ""),
        )
        wb.save(EXCEL_FILE)

    _with_lock_sync(job)


def _upsert_fac(wb, f, is_upd: bool):
    ws = wb["Factures"]
    tot = float(f.get("montant_total") or 0)
    pay = float(f.get("montant_regle") or 0)
    rest = tot - pay
    st = f.get("statut") or "Impayee"
    pat = f"{f.get('patient_nom') or ''} {f.get('patient_prenom') or ''}".strip()
    row_data = [
        f.get("id_facture"),
        f.get("id_patient"),
        pat,
        f.get("numero_facture") or "",
        fd(f.get("date_facture")),
        f.get("motif") or "",
        f.get("dent") or "",
        tot,
        pay,
        rest,
        st,
        fdt(datetime.now()),
    ]
    st_c = T["Factures"].get("st", {}).get(st, {"f": "888888", "b": "EEEEEE"})

    exist_rn = None
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == f.get("id_facture"):
            exist_rn = row
            break

    if exist_rn:
        for i, v in enumerate(row_data, 1):
            ws.cell(row=exist_rn, column=i, value=v)
        for c in range(1, 13):
            cell = ws.cell(row=exist_rn, column=c)
            cell.fill = PatternFill("solid", fgColor=_argb(st_c["b"]))
            cell.font = Font(size=10, name="Calibri")
            cell.alignment = Alignment(vertical="center")
        ws.cell(row=exist_rn, column=11).font = Font(bold=True, color=_argb(st_c["f"]), size=10, name="Calibri")
        for c in (8, 9, 10):
            ws.cell(row=exist_rn, column=c).number_format = '#,##0.00 "DH"'
    else:
        rn = ws.max_row + 1
        ws.append(row_data)
        _style_data_row(ws, rn, "Factures", st_c["b"])
        ws.cell(row=rn, column=11).font = Font(bold=True, color=_argb(st_c["f"]), size=10, name="Calibri")
        for c in (8, 9, 10):
            ws.cell(row=rn, column=c).number_format = '#,##0.00 "DH"'

    typ = "FACTURE_MAJ" if is_upd else "FACTURE_AJOUT"
    det = (
        f"{f.get('numero_facture')} | Payé: {pay:.2f} DH | Reste: {rest:.2f} DH | {st}"
        if is_upd
        else f"{f.get('numero_facture')} | Total: {tot:.2f} DH | {f.get('motif') or '-'}"
    )
    _hist(wb, typ, "Facture mise à jour" if is_upd else "Facture créée", pat, det)


def append_facture(facture_id: int):
    rows = _q(
        "SELECT f.*, pt.nom AS patient_nom, pt.prenom AS patient_prenom FROM facture f "
        "LEFT JOIN patient pt ON f.id_patient=pt.id_patient WHERE f.id_facture=%s",
        (facture_id,),
    )
    if not rows:
        return

    def job():
        wb = _get_wb()
        _upsert_fac(wb, rows[0], False)
        wb.save(EXCEL_FILE)

    _with_lock_sync(job)


def append_facture_update(facture_id):
    if not facture_id:
        return
    rows = _q(
        "SELECT f.*, pt.nom AS patient_nom, pt.prenom AS patient_prenom FROM facture f "
        "LEFT JOIN patient pt ON f.id_patient=pt.id_patient WHERE f.id_facture=%s",
        (facture_id,),
    )
    if not rows:
        return

    def job():
        wb = _get_wb()
        _upsert_fac(wb, rows[0], True)
        wb.save(EXCEL_FILE)

    _with_lock_sync(job)


def fire_excel(fn, *args):
    threading.Thread(target=lambda: fn(*args), daemon=True).start()
