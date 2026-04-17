"""
DentiPro — API Flask (équivalent Express server.js + routes).
Point d'entrée: python app.py  → http://localhost:3000
"""
from __future__ import annotations

import io
import os
import random
import re
from datetime import date, datetime, timedelta
from decimal import Decimal
from functools import wraps
from pathlib import Path

import jwt
import bcrypt
import pymysql
from flask import Flask, jsonify, request, send_file, send_from_directory, g
from flask_cors import CORS
from pymysql.cursors import DictCursor
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from werkzeug.utils import secure_filename

import excel_sync
from extras import ensure_extra_tables, register_extras_routes

PROJECT_ROOT = Path(__file__).resolve().parent.parent
FRONTEND_DIR = PROJECT_ROOT / "frontend"
UPLOAD_IMAG = PROJECT_ROOT / "uploads" / "imagerie"
UPLOAD_CHAT = PROJECT_ROOT / "uploads" / "chat"

SECRET = os.environ.get("JWT_SECRET", "cabinet_dentaire_secret_key_2024")
MYSQL = dict(
    host=os.environ.get("MYSQL_HOST", "localhost"),
    user=os.environ.get("MYSQL_USER", "root"),
    password=os.environ.get("MYSQL_PASSWORD", ""),
    database=os.environ.get("MYSQL_DATABASE", "cabinet_dentaire"),
    charset="utf8mb4",
    cursorclass=DictCursor,
)


def _conn():
    return pymysql.connect(**MYSQL)


def query_all(sql: str, params=None):
    conn = _conn()
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params or ())
            return cur.fetchall()
    except Exception as e:
        print(f"❌ Erreur SQL (query_all): {e}")
        return []
    finally:
        conn.close()

def execute(sql: str, params=None, fetch_last=False):
    conn = _conn()
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params or ())
            conn.commit()
            return cur.lastrowid if fetch_last else None
    except Exception as e:
        conn.rollback()
        print(f"❌ Erreur SQL (execute): {e}")
        raise
    finally:
        conn.close()


def executemany(sql: str, seq):
    if not seq:
        return
    conn = _conn()
    try:
        with conn.cursor() as cur:
            cur.executemany(sql, seq)
            conn.commit()
    finally:
        conn.close()


excel_sync.configure(query_all)


def _san(v):
    if v is None:
        return None
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, datetime):
        return v.isoformat(sep=" ")
    if isinstance(v, date):
        return str(v)
    if isinstance(v, timedelta):
        secs = int(v.total_seconds()) % 86400
        h, r = divmod(secs, 3600)
        m, s = divmod(r, 60)
        return f"{h:02d}:{m:02d}:{s:02d}"
    return v


def sj(row):
    return {k: _san(v) for k, v in row.items()}


def sjl(rows):
    return [sj(r) for r in rows]


def _today_local():
    n = datetime.now()
    return f"{n.year}-{n.month:02d}-{n.day:02d}"


def _month_bounds():
    n = datetime.now()
    first = date(n.year, n.month, 1)
    if n.month == 12:
        last = date(n.year, 12, 31)
    else:
        last = date(n.year, n.month + 1, 1) - timedelta(days=1)
    return str(first), str(last)


def _auth_token():
    h = request.headers.get("Authorization") or ""
    parts = h.split()
    # Si le header est "Bearer TOKEN", on prend le TOKEN
    if len(parts) > 1:
        return parts[1]
    # Sinon, on prend tout ce qui vient (cas où le JS envoie juste le TOKEN)
    return h if h else None

def require_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = _auth_token()
        if not token:
            return jsonify({"error": "Accès non autorisé"}), 401
        try:
            g.user = jwt.decode(token, SECRET, algorithms=["HS256"])
        except jwt.PyJWTError:
            return jsonify({"error": "Token invalide ou expiré"}), 401
        return f(*args, **kwargs)
    return decorated

def create_app():
    app = Flask(__name__)
    CORS(app)
    try:
        ensure_extra_tables(_conn)
    except Exception as ex:
        print("ensure_extra_tables:", ex)
    UPLOAD_IMAG.mkdir(parents=True, exist_ok=True)
    UPLOAD_CHAT.mkdir(parents=True, exist_ok=True)

    # ── Auth ─────────────────────────────────────────────
    @app.post("/api/login")
    def login():
        b = request.get_json(silent=True) or {}
        u, pw = b.get("username"), b.get("password")
        rows = query_all(
            "SELECT * FROM utilisateur WHERE username=%s", (u,)
        )
        if not rows:
            return jsonify({"success": False, "message": "Identifiants incorrects"})
        user = rows[0]
        
        # Vérification du mot de passe (fallback texte brut pour les utilisateurs existants)
        db_pwd = user["password"]
        try:
            is_valid = bcrypt.checkpw(pw.encode('utf-8'), db_pwd.encode('utf-8'))
        except ValueError:
            is_valid = (pw == db_pwd)
            
        if not is_valid:
            return jsonify({"success": False, "message": "Identifiants incorrects"})

        token = jwt.encode(
            {
                "id": user["id_user"],
                "username": user["username"],
                "role": user["role"],
            },
            SECRET,
            algorithm="HS256",
        )
        if isinstance(token, bytes):
            token = token.decode()
        lid = user.get("linked_user_id")
        return jsonify(
            {
                "success": True,
                "token": token,
                "user": {
                    "id": user["id_user"],
                    "username": user["username"],
                    "nom": user["nom"],
                    "prenom": user["prenom"],
                    "role": user["role"],
                    "email": user.get("email") or "",
                    "linked_user_id": int(lid) if lid is not None else None,
                },
            }
        )

    @app.get("/api/verify")
    def verify():
        t = _auth_token()
        if not t:
            return jsonify({"valid": False}), 401
        try:
            payload = jwt.decode(t, SECRET, algorithms=["HS256"])
            return jsonify({"valid": True, "user": payload})
        except jwt.PyJWTError:
            return jsonify({"valid": False}), 401

    @app.put("/api/change-password")
    @require_auth
    def change_password():
        b = request.get_json(silent=True) or {}
        uid = g.user["id"]  # SECURISE : Récupère l'ID depuis le token
        old_p, new_p = b.get("old_password"), b.get("new_password")
        r = query_all(
            "SELECT id_user, password FROM utilisateur WHERE id_user=%s",
            (uid,),
        )
        if not r:
            return jsonify({"error": "Utilisateur non trouvé"}), 404
            
        db_pwd = r[0]["password"]
        try:
            is_valid = bcrypt.checkpw(old_p.encode('utf-8'), db_pwd.encode('utf-8'))
        except ValueError:
            is_valid = (old_p == db_pwd)
            
        if not is_valid:
            return jsonify({"error": "Mot de passe actuel incorrect"}), 400
            
        hashed_p = bcrypt.hashpw(new_p.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        execute(
            "UPDATE utilisateur SET password=%s WHERE id_user=%s", (hashed_p, uid)
        )
        return jsonify({"success": True})

    @app.put("/api/profile")
    @require_auth
    def profile():
        b = request.get_json(silent=True) or {}
        uid = g.user["id"]  # SECURISE : Récupère l'ID depuis le token
        execute(
            "UPDATE utilisateur SET nom=%s,prenom=%s,email=%s WHERE id_user=%s",
            (b.get("nom"), b.get("prenom"), b.get("email") or None, uid),
        )
        return jsonify({"success": True})

    @app.post("/api/logout")
    def logout():
        return jsonify({"success": True})

    # ── Stats (server.js) ─────────────────────────────────
    @app.get("/api/paiements/stats/types")
    @require_auth
    def pay_types():
        r = query_all(
            """SELECT
            COALESCE(SUM(CASE WHEN type_paiement='Especes'  THEN montant ELSE 0 END),0) as especes,
            COALESCE(SUM(CASE WHEN type_paiement='Carte'    THEN montant ELSE 0 END),0) as carte,
            COALESCE(SUM(CASE WHEN type_paiement='Virement' THEN montant ELSE 0 END),0) as virement,
            COALESCE(SUM(CASE WHEN type_paiement='Cheque'   THEN montant ELSE 0 END),0) as cheque
            FROM paiement"""
        )
        return jsonify(sj(r[0]) if r else {})

    @app.get("/api/paiements/stats/chart")
    @require_auth
    def pay_chart():
        r = query_all(
            """SELECT DATE_FORMAT(date_paiement,'%%Y-%%m') as mois, SUM(montant) as total
            FROM paiement WHERE date_paiement >= DATE_SUB(NOW(), INTERVAL 12 MONTH)
            GROUP BY mois ORDER BY mois ASC"""
        )
        return jsonify(sjl(r))

    @app.get("/api/rdv/stats/chart")
    @require_auth
    def rdv_chart():
        r = query_all(
            """SELECT DATE_FORMAT(date_rdv,'%%Y-%%m') as mois, COUNT(*) as total
            FROM rendez_vous WHERE date_rdv >= DATE_SUB(NOW(), INTERVAL 12 MONTH)
            GROUP BY mois ORDER BY mois ASC"""
        )
        return jsonify(sjl(r))

    register_patients_routes(app)
    register_rdv_routes(app)
    register_paiements_routes(app)
    register_factures_routes(app)
    register_salle_routes(app)
    register_schema_routes(app)
    register_chat_routes(app)
    register_export_routes(app)
    register_extras_routes(app, query_all, execute, sj, sjl)

    @app.route("/uploads/<path:subpath>")
    def uploads_static(subpath):
        return send_from_directory(PROJECT_ROOT / "uploads", subpath)

    @app.route("/")
    def root():
        return send_from_directory(FRONTEND_DIR, "login.html")

    @app.route("/index")
    def index_dash():
        return send_from_directory(FRONTEND_DIR, "dashboard.html")

    @app.route("/<path:fname>")
    def frontend_static(fname):
        if fname.startswith("api/"):
            return jsonify({"error": "Not found"}), 404
        fp = FRONTEND_DIR / fname
        if fp.is_file():
            return send_from_directory(FRONTEND_DIR, fname)
        return jsonify({"error": "Not found"}), 404

    return app


def register_patients_routes(app):
    @app.get("/api/patients/stats/month")
    @require_auth
    def p_stats_month():
        first, last = _month_bounds()
        r = query_all(
            "SELECT COUNT(*) as total_patients FROM patient WHERE DATE(created_at) BETWEEN %s AND %s",
            (first, last),
        )
        return jsonify(sj(r[0]))

    @app.get("/api/patients/stats/chart")
    @require_auth
    def p_stats_chart():
        r = query_all(
            "SELECT DATE_FORMAT(created_at,'%%Y-%%m') as mois, COUNT(*) as total FROM patient "
            "WHERE created_at >= DATE_SUB(NOW(), INTERVAL 12 MONTH) GROUP BY mois ORDER BY mois ASC"
        )
        return jsonify(sjl(r))

    @app.get("/api/patients/search/<path:query>")
    @require_auth
    def p_search(query):
        q = f"%{query}%"
        r = query_all(
            "SELECT * FROM patient WHERE nom LIKE %s OR prenom LIKE %s OR cnie LIKE %s OR telephone LIKE %s ORDER BY nom ASC",
            (q, q, q, q),
        )
        return jsonify(sjl(r))

    @app.delete("/api/patients/imagerie/<int:img_id>")
    @require_auth
    def p_img_del(img_id):
        rows = query_all("SELECT fichier_url FROM patient_imagerie WHERE id = %s", (img_id,))
        if not rows:
            return jsonify({"error": "Non trouvé"}), 404
        rel = rows[0]["fichier_url"].lstrip("/").replace("/", os.sep)
        fp = PROJECT_ROOT / rel
        try:
            if fp.is_file():
                fp.unlink()
        except OSError:
            pass
        execute("DELETE FROM patient_imagerie WHERE id = %s", (img_id,))
        return jsonify({"success": True})

    @app.get("/api/patients")
    @require_auth
    def p_list():
        r = query_all("SELECT * FROM patient ORDER BY nom ASC")
        return jsonify(sjl(r))

    @app.post("/api/patients")
    @require_auth
    def p_add():
        b = request.get_json(silent=True) or {}
        sql = """INSERT INTO patient
            (nom,prenom,sexe,telephone,cnie,date_naissance,ville,pays,email,
             type_assurance,numero_immatriculation,type_patient,
             antecedents_medicaux,allergies,notes)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
        pid = execute(
            sql,
            (
                b.get("nom"),
                b.get("prenom"),
                b.get("sexe"),
                b.get("telephone"),
                b.get("cnie") or None,
                b.get("date") or None,
                b.get("ville") or None,
                b.get("pays") or "Maroc",
                b.get("email") or None,
                b.get("type_assurance") or "Aucune",
                b.get("numero_immatriculation") or None,
                b.get("type_patient") or "adulte",
                b.get("antecedents_medicaux") or None,
                b.get("allergies") or None,
                b.get("notes") or None,
            ),
            fetch_last=True,
        )
        excel_sync.fire_excel(excel_sync.append_patient, pid)
        return jsonify({"success": True, "id": pid})

    @app.post("/api/patients/<int:pid>/imagerie")
    @require_auth
    def p_img_up(pid):
        if "fichier" not in request.files:
            return (
                jsonify(
                    {"error": "Aucun fichier reçu ou fichier trop volumineux (max 50MB)"}
                ),
                400,
            )
        f = request.files["fichier"]
        if not f.filename:
            return (
                jsonify(
                    {"error": "Aucun fichier reçu ou fichier trop volumineux (max 50MB)"}
                ),
                400,
            )
        ALLOWED_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.pdf', '.doc', '.docx'}
        ext = Path(f.filename).suffix.lower()
        if ext not in ALLOWED_EXTENSIONS:
            return jsonify({"error": "Type de fichier non autorisé."}), 400
        raw = secure_filename(f.filename) or "file"
        safe = re.sub(r"[^a-zA-Z0-9._-]", "_", f.filename)
        name = f"{int(datetime.now().timestamp() * 1000)}_{safe}"
        path = UPLOAD_IMAG / name
        f.save(path)
        if path.stat().st_size > 50 * 1024 * 1024:
            path.unlink(missing_ok=True)
            return (
                jsonify(
                    {"error": "Aucun fichier reçu ou fichier trop volumineux (max 50MB)"}
                ),
                400,
            )
        b = request.form
        fichier_url = f"/uploads/imagerie/{name}"
        iid = execute(
            """INSERT INTO patient_imagerie
            (id_patient,id_user,type_doc,titre,fichier_url,fichier_nom,fichier_type,description,date_doc)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (
                pid,
                g.user["id"],  # SECURISE: Utilise l'ID de l'utilisateur authentifié
                b.get("type_doc") or "scanner",
                b.get("titre") or f.filename,
                fichier_url,
                f.filename,
                f.mimetype or "application/octet-stream",
                b.get("description") or None,
                b.get("date_doc") or None,
            ),
            fetch_last=True,
        )
        return jsonify(
            {"success": True, "id": iid, "url": fichier_url, "nom": f.filename}
        )

    @app.get("/api/patients/<int:pid>/imagerie")
    @require_auth
    def p_img_list(pid):
        r = query_all(
            "SELECT * FROM patient_imagerie WHERE id_patient = %s ORDER BY created_at DESC",
            (pid,),
        )
        return jsonify(sjl(r))

    @app.get("/api/patients/<int:pid>")
    @require_auth
    def p_one(pid):
        r = query_all("SELECT * FROM patient WHERE id_patient = %s", (pid,))
        if not r:
            return jsonify({"error": "Patient non trouvé"}), 404
        patient = sj(r[0])
        st = query_all(
            """SELECT COALESCE(SUM(f.montant_total),0) AS total_facture,
            COALESCE(SUM(f.montant_regle),0) AS total_regle,
            COALESCE(SUM(f.montant_total - f.montant_regle),0) AS total_restant
            FROM facture f WHERE f.id_patient = %s""",
            (pid,),
        )
        if st:
            patient["stats"] = sj(st[0])
        lr = query_all(
            "SELECT date_rdv,motif FROM rendez_vous WHERE id_patient=%s ORDER BY date_rdv DESC LIMIT 1",
            (pid,),
        )
        patient["dernier_rdv"] = sj(lr[0]) if lr else None
        now = _today_local()
        nr = query_all(
            "SELECT date_rdv,heure_rdv,motif FROM rendez_vous WHERE id_patient=%s AND date_rdv>=%s AND statut='Prevu' ORDER BY date_rdv ASC,heure_rdv ASC LIMIT 1",
            (pid, now),
        )
        patient["prochain_rdv"] = sj(nr[0]) if nr else None
        return jsonify(patient)

    @app.put("/api/patients/<int:pid>")
    @require_auth
    def p_up(pid):
        b = request.get_json(silent=True) or {}
        execute(
            """UPDATE patient SET
            nom=%s,prenom=%s,sexe=%s,telephone=%s,cnie=%s,date_naissance=%s,
            ville=%s,pays=%s,email=%s,type_assurance=%s,numero_immatriculation=%s,type_patient=%s,
            antecedents_medicaux=%s,allergies=%s,notes=%s
            WHERE id_patient=%s""",
            (
                b.get("nom"),
                b.get("prenom"),
                b.get("sexe"),
                b.get("telephone"),
                b.get("cnie") or None,
                b.get("date") or None,
                b.get("ville") or None,
                b.get("pays") or "Maroc",
                b.get("email") or None,
                b.get("type_assurance") or "Aucune",
                b.get("numero_immatriculation") or None,
                b.get("type_patient") or "adulte",
                b.get("antecedents_medicaux") or None,
                b.get("allergies") or None,
                b.get("notes") or None,
                pid,
            ),
        )
        return jsonify({"success": True})

    @app.delete("/api/patients/<int:pid>")
    @require_auth
    def p_del(pid):
        execute("DELETE FROM patient WHERE id_patient = %s", (pid,))
        return jsonify({"success": True})


def register_rdv_routes(app):
    @app.get("/api/rdv")
    @require_auth
    def rdv_list():
        d = request.args.get("date")
        pid = request.args.get("patient_id")
        sql = """
            SELECT r.*, p.nom, p.prenom
            FROM rendez_vous r
            JOIN patient p ON r.id_patient = p.id_patient
            WHERE 1=1
        """
        params = []
        if d:
            sql += " AND r.date_rdv = %s"
            params.append(d)
        if pid:
            sql += " AND r.id_patient = %s"
            params.append(pid)
        sql += " ORDER BY r.date_rdv DESC, r.heure_rdv DESC"
        return jsonify(sjl(query_all(sql, tuple(params))))

    @app.get("/api/rdv/today")
    @require_auth
    def rdv_today():
        t = _today_local()
        sql = """
            SELECT r.*, p.nom, p.prenom
            FROM rendez_vous r
            JOIN patient p ON r.id_patient = p.id_patient
            WHERE r.date_rdv = %s
            ORDER BY r.heure_rdv ASC
        """
        return jsonify(sjl(query_all(sql, (t,))))

    @app.get("/api/rdv/stats/count")
    @require_auth
    def rdv_stats_count():
        t = _today_local()
        r = query_all(
            """SELECT COUNT(*) as total,
            SUM(CASE WHEN date_rdv = %s THEN 1 ELSE 0 END) as today
            FROM rendez_vous""",
            (t,),
        )
        return jsonify(sj(r[0]))

    @app.get("/api/rdv/waiting-room")
    @require_auth
    def rdv_wait():
        t = _today_local()
        sql = """
            SELECT r.*, p.nom, p.prenom
            FROM rendez_vous r
            JOIN patient p ON r.id_patient = p.id_patient
            WHERE r.date_rdv = %s AND r.statut != 'Annule'
            ORDER BY r.heure_rdv ASC
        """
        rows = query_all(sql, (t,))
        out = []
        for i, row in enumerate(rows):
            d = sj(row)
            d["position"] = i + 1
            out.append(d)
        return jsonify(out)

    @app.get("/api/rdv/<int:rid>")
    @require_auth
    def rdv_one(rid):
        sql = """
            SELECT r.*, p.nom, p.prenom
            FROM rendez_vous r
            JOIN patient p ON r.id_patient = p.id_patient
            WHERE r.id_rdv = %s
        """
        r = query_all(sql, (rid,))
        if not r:
            return jsonify({"error": "Rendez-vous non trouvé"}), 404
        return jsonify(sj(r[0]))

    @app.post("/api/rdv")
    @require_auth
    def rdv_add():
        b = request.get_json(silent=True) or {}
        iid = execute(
            """INSERT INTO rendez_vous (id_patient, date_rdv, heure_rdv, statut, motif, dent, id_user)
            VALUES (%s,%s,%s,%s,%s,%s,%s)""",
            (
                b.get("id_patient"),
                b.get("date_rdv"),
                b.get("heure_rdv"),
                b.get("statut") or "Prevu",
                b.get("motif"),
                b.get("dent") or None,
                g.user["id"],  # SECURISE: Utilise l'ID de l'utilisateur authentifié
            ),
            fetch_last=True,
        )
        excel_sync.fire_excel(excel_sync.append_rdv, iid)
        return jsonify({"success": True, "id": iid})

    @app.put("/api/rdv/<int:rid>")
    @require_auth
    def rdv_put(rid):
        b = request.get_json(silent=True) or {}
        execute(
            """UPDATE rendez_vous
            SET id_patient=%s, date_rdv=%s, heure_rdv=%s, statut=%s, motif=%s, dent=%s
            WHERE id_rdv=%s""",
            (
                b.get("id_patient"),
                b.get("date_rdv"),
                b.get("heure_rdv"),
                b.get("statut"),
                b.get("motif"),
                b.get("dent") or None,
                rid,
            ),
        )
        return jsonify({"success": True})

    @app.patch("/api/rdv/<int:rid>/statut")
    @require_auth
    def rdv_statut(rid):
        b = request.get_json(silent=True) or {}
        execute("UPDATE rendez_vous SET statut = %s WHERE id_rdv = %s", (b.get("statut"), rid))
        return jsonify({"success": True})

    @app.delete("/api/rdv/<int:rid>")
    @require_auth
    def rdv_del(rid):
        execute("DELETE FROM rendez_vous WHERE id_rdv = %s", (rid,))
        return jsonify({"success": True})

    @app.patch("/api/rdv/<int:rid>/en-cours")
    @require_auth
    def rdv_encours(rid):
        execute("UPDATE rendez_vous SET statut = 'En cours' WHERE id_rdv = %s", (rid,))
        return jsonify({"success": True})


def _update_facture_montant(id_facture):
    if not id_facture:
        return
    pay = query_all(
        "SELECT COALESCE(SUM(montant), 0) as total_paye FROM paiement WHERE id_facture = %s",
        (id_facture,),
    )
    new_regle = float(pay[0]["total_paye"] if pay else 0)
    fac = query_all("SELECT montant_total FROM facture WHERE id_facture = %s", (id_facture,))
    if not fac:
        return
    mt = float(fac[0]["montant_total"])
    st = "Impayee"
    if new_regle >= mt:
        st = "Payee"
    elif new_regle > 0:
        st = "Partiellement payee"
    execute(
        "UPDATE facture SET montant_regle = %s, statut = %s WHERE id_facture = %s",
        (new_regle, st, id_facture),
    )


def register_paiements_routes(app):
    @app.get("/api/paiements/stats/summary")
    def pay_sum():
        dd, df = request.args.get("date_debut"), request.args.get("date_fin")
        where, params = "", []
        if dd and df:
            where, params = "WHERE date_paiement BETWEEN %s AND %s", [dd, df]
        elif dd:
            where, params = "WHERE date_paiement >= %s", [dd]
        elif df:
            where, params = "WHERE date_paiement <= %s", [df]
        sql = f"""SELECT COUNT(*) as total_paiements, SUM(montant) as total_montant,
        SUM(CASE WHEN type_paiement='Especes' THEN montant ELSE 0 END) as especes,
        SUM(CASE WHEN type_paiement='Carte'   THEN montant ELSE 0 END) as carte,
        SUM(CASE WHEN type_paiement='Virement' THEN montant ELSE 0 END) as virement,
        SUM(CASE WHEN type_paiement='Cheque'  THEN montant ELSE 0 END) as cheque
        FROM paiement {where}"""
        r = query_all(sql, tuple(params))
        return jsonify(sj(r[0]))

    @app.get("/api/paiements/stats/month")
    def pay_month():
        n = datetime.now()
        first = date(n.year, n.month, 1).isoformat()
        if n.month == 12:
            last = date(n.year, 12, 31).isoformat()
        else:
            last = (date(n.year, n.month + 1, 1) - timedelta(days=1)).isoformat()
        r = query_all(
            "SELECT COALESCE(SUM(montant), 0) as total_montant FROM paiement WHERE date_paiement BETWEEN %s AND %s",
            (first, last),
        )
        return jsonify(sj(r[0]))

    @app.get("/api/paiements/stats/today")
    def pay_today():
        today = datetime.utcnow().date().isoformat()
        r = query_all(
            "SELECT COALESCE(SUM(montant), 0) as total_montant FROM paiement WHERE date_paiement = %s",
            (today,),
        )
        return jsonify(sj(r[0]))

    @app.get("/api/paiements/patient/<int:patient_id>/total")
    def pay_ptotal(patient_id):
        r = query_all(
            "SELECT COALESCE(SUM(montant), 0) as total_paye FROM paiement WHERE id_patient = %s",
            (patient_id,),
        )
        return jsonify(sj(r[0]))

    @app.get("/api/paiements/patient/<int:patient_id>")
    def pay_plist(patient_id):
        sql = """
        SELECT p.*, pat.nom as patient_nom, pat.prenom as patient_prenom, r.motif as rdv_motif
        FROM paiement p
        LEFT JOIN patient pat ON p.id_patient = pat.id_patient
        LEFT JOIN rendez_vous r ON p.id_rdv = r.id_rdv
        WHERE p.id_patient = %s
        ORDER BY p.date_paiement DESC
        """
        return jsonify(sjl(query_all(sql, (patient_id,))))

    @app.get("/api/paiements")
    def pay_list():
        pid = request.args.get("patient_id")
        dd = request.args.get("date_debut")
        df = request.args.get("date_fin")
        fid = request.args.get("id_facture")
        sql = """
        SELECT p.*, pat.nom as patient_nom, pat.prenom as patient_prenom, r.motif as rdv_motif,
               f.montant_total as facture_total, f.montant_regle as facture_regle, f.numero_facture
        FROM paiement p
        LEFT JOIN patient pat ON p.id_patient = pat.id_patient
        LEFT JOIN rendez_vous r ON p.id_rdv = r.id_rdv
        LEFT JOIN facture f ON p.id_facture = f.id_facture
        WHERE 1=1
        """
        params = []
        if pid:
            sql += " AND p.id_patient = %s"
            params.append(pid)
        if dd:
            sql += " AND p.date_paiement >= %s"
            params.append(dd)
        if df:
            sql += " AND p.date_paiement <= %s"
            params.append(df)
        if fid:
            sql += " AND p.id_facture = %s"
            params.append(fid)
        sql += " ORDER BY p.date_paiement DESC, p.id_paiement DESC"
        return jsonify(sjl(query_all(sql, tuple(params))))

    @app.get("/api/paiements/<int:pay_id>")
    def pay_one(pay_id):
        sql = """
        SELECT p.*, pat.nom as patient_nom, pat.prenom as patient_prenom, r.motif as rdv_motif
        FROM paiement p
        LEFT JOIN patient pat ON p.id_patient = pat.id_patient
        LEFT JOIN rendez_vous r ON p.id_rdv = r.id_rdv
        WHERE p.id_paiement = %s
        """
        r = query_all(sql, (pay_id,))
        if not r:
            return jsonify({"error": "Paiement non trouvé"}), 404
        return jsonify(sj(r[0]))

    @app.post("/api/paiements")
    def pay_add():
        b = request.get_json(silent=True) or {}
        iid = execute(
            """INSERT INTO paiement (id_patient, id_rdv, id_facture, id_user, montant, type_paiement, date_paiement, notes)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
            (
                b.get("id_patient"),
                b.get("id_rdv") or None,
                b.get("id_facture") or None,
                b.get("id_user"),
                b.get("montant"),
                b.get("type_paiement") or "Especes",
                b.get("date_paiement"),
                b.get("notes") or None,
            ),
            fetch_last=True,
        )
        excel_sync.fire_excel(excel_sync.append_paiement, iid)
        fid = b.get("id_facture")
        if fid:
            _update_facture_montant(fid)
            excel_sync.fire_excel(excel_sync.append_facture_update, fid)
        return jsonify({"success": True, "id": iid})

    @app.put("/api/paiements/<int:pay_id>")
    def pay_put(pay_id):
        b = request.get_json(silent=True) or {}
        old = query_all("SELECT id_facture FROM paiement WHERE id_paiement = %s", (pay_id,))
        old_f = old[0]["id_facture"] if old else None
        execute(
            """UPDATE paiement SET id_patient=%s,id_rdv=%s,id_facture=%s,montant=%s,type_paiement=%s,date_paiement=%s,notes=%s WHERE id_paiement=%s""",
            (
                b.get("id_patient"),
                b.get("id_rdv") or None,
                b.get("id_facture") or None,
                b.get("montant"),
                b.get("type_paiement"),
                b.get("date_paiement"),
                b.get("notes") or None,
                pay_id,
            ),
        )
        nf = b.get("id_facture") or None
        tasks = []
        if old_f and old_f != nf:
            tasks.append(old_f)
        if nf:
            tasks.append(nf)
        for fid in tasks:
            _update_facture_montant(fid)
            excel_sync.fire_excel(excel_sync.append_facture_update, fid)
        return jsonify({"success": True})

    @app.delete("/api/paiements/<int:pay_id>")
    def pay_del(pay_id):
        old = query_all("SELECT id_facture FROM paiement WHERE id_paiement = %s", (pay_id,))
        old_f = old[0]["id_facture"] if old else None
        execute("DELETE FROM paiement WHERE id_paiement = %s", (pay_id,))
        if old_f:
            _update_facture_montant(old_f)
            excel_sync.fire_excel(excel_sync.append_facture_update, old_f)
        return jsonify({"success": True})


def _pdf_facture(facture, details, paiements):
    """Generate a fully styled A4 invoice PDF matching the original Node/PDFKit version."""
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4
    blue   = colors.HexColor("#1a73e8")
    green  = colors.HexColor("#2e7d32")
    red    = colors.HexColor("#c62828")
    orange = colors.HexColor("#e65100")
    grey_bg= colors.HexColor("#f0f4ff")
    light_green = colors.HexColor("#e8f5e9")
    light_red   = colors.HexColor("#ffebee")
    light_orange= colors.HexColor("#fff3e0")
    white  = colors.white
    black  = colors.black

    def fmt_date(v):
        if not v:
            return "-"
        if isinstance(v, datetime):
            return v.strftime("%d/%m/%Y")
        if isinstance(v, date):
            return v.strftime("%d/%m/%Y")
        try:
            return datetime.strptime(str(v)[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            return str(v)[:10]

    def draw_rect_filled(x, y, rw, rh, fill_color, stroke_color=None):
        c.setFillColor(fill_color)
        if stroke_color:
            c.setStrokeColor(stroke_color)
            c.rect(x, y, rw, rh, fill=1, stroke=1)
        else:
            c.rect(x, y, rw, rh, fill=1, stroke=0)
        c.setStrokeColor(black)

    # ── HEADER BAND ─────────────────────────────────────────────
    draw_rect_filled(0, h - 80, w, 80, blue)
    c.setFillColor(white)
    c.setFont("Helvetica-Bold", 26)
    c.drawCentredString(w / 2, h - 30, "CABINET DENTAIRE")
    c.setFont("Helvetica", 11)
    c.drawCentredString(w / 2, h - 50, "Système de Gestion Professionnel")
    c.setFillColor(black)

    y = h - 100

    mt = float(facture.get("montant_total") or 0)
    mr = float(facture.get("montant_regle") or 0)
    rest = mt - mr

    # ── INVOICE INFO BOX (right) ─────────────────────────────────
    draw_rect_filled(370, y, 175, 72, grey_bg, blue)
    c.setFillColor(blue)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(378, y + 56, "FACTURE")
    c.setFillColor(colors.HexColor("#333333"))
    c.setFont("Helvetica", 10)
    c.drawString(378, y + 40, f"N°: {facture.get('numero_facture') or '-'}")
    c.drawString(378, y + 26, f"Date: {fmt_date(facture.get('date_facture'))}")
    c.drawString(378, y + 12, f"Réf: FAC-{facture.get('id_facture') or '-'}")
    c.setFillColor(black)

    # ── PATIENT INFO (left) ──────────────────────────────────────
    c.setFont("Helvetica-Bold", 12)
    c.drawString(50, y, "INFORMATIONS PATIENT")
    c.setFillColor(blue)
    c.setLineWidth(1.5)
    c.line(50, y - 4, 250, y - 4)
    c.setLineWidth(1)
    c.setFillColor(colors.HexColor("#444444"))
    c.setFont("Helvetica", 10)
    y -= 20
    c.drawString(50, y, f"Nom: {facture.get('patient_nom') or ''} {facture.get('patient_prenom') or ''}")
    y -= 14
    c.drawString(50, y, f"Téléphone: {facture.get('patient_telephone') or '-'}")
    if facture.get("patient_cnie"):
        y -= 14
        c.drawString(50, y, f"CIN: {facture['patient_cnie']}")
    c.setFillColor(black)

    y = max(y, h - 183) - 16

    # ── MOTIF / DENT BAND ────────────────────────────────────────
    if facture.get("motif") or facture.get("dent"):
        draw_rect_filled(50, y - 6, 495, 34, colors.HexColor("#e8f0fe"))
        c.setFillColor(blue)
        c.setFont("Helvetica-Bold", 10)
        acte = ""
        if facture.get("motif"):
            acte += f"Motif: {facture['motif']}"
        if facture.get("dent"):
            acte += f"   |   Dent(s): {facture['dent']}"
        c.drawString(58, y + 6, acte)
        c.setFillColor(black)
        y -= 44

    # ── TABLE HEADER ─────────────────────────────────────────────
    col_x = [50, 290, 340, 435]
    col_w = [240, 50, 95, 95]
    headers = ["Description", "Qté", "Prix Unitaire", "Sous-total"]
    draw_rect_filled(50, y - 4, 480, 26, blue)
    c.setFillColor(white)
    c.setFont("Helvetica-Bold", 10)
    for i, hdr in enumerate(headers):
        align = "left" if i == 0 else "right"
        if align == "right":
            c.drawRightString(col_x[i] + col_w[i], y + 6, hdr)
        else:
            c.drawString(col_x[i], y + 6, hdr)
    c.setFillColor(black)
    y -= 4 + 22

    # ── TABLE ROWS ───────────────────────────────────────────────
    c.setFont("Helvetica", 9)
    even_bg = colors.HexColor("#f8f9fa")
    if not details:
        desc = facture.get("motif") or "Consultation dentaire"
        if facture.get("dent"):
            desc += f" — Dent(s) {facture['dent']}"
        draw_rect_filled(50, y - 4, 480, 20, even_bg)
        c.setFillColor(colors.HexColor("#333333"))
        c.drawString(col_x[0], y + 2, desc[:70])
        c.drawRightString(col_x[1] + col_w[1], y + 2, "1")
        c.drawRightString(col_x[2] + col_w[2], y + 2, f"{mt:.2f} DH")
        c.drawRightString(col_x[3] + col_w[3], y + 2, f"{mt:.2f} DH")
        c.setFillColor(black)
        y -= 20
    else:
        for idx, d in enumerate(details):
            if y < 120:
                # Footer on current page
                draw_rect_filled(0, 0, w, 40, blue)
                c.setFillColor(white)
                c.setFont("Helvetica", 8)
                c.drawCentredString(w / 2, 15, "Merci pour votre confiance ! — Généré par DentiPro")
                c.showPage()
                y = h - 50
                c.setFont("Helvetica", 9)
            row_bg = even_bg if idx % 2 == 0 else white
            draw_rect_filled(50, y - 4, 480, 20, row_bg)
            q  = int(d.get("quantite") or 1)
            pu = float(d.get("prix_unitaire") or 0)
            st_val = float(d.get("sous_total") or 0)
            desc = (d.get("description") or d.get("nom_service") or "Prestation")[:60]
            c.setFillColor(colors.HexColor("#333333"))
            c.drawString(col_x[0], y + 2, desc)
            c.drawRightString(col_x[1] + col_w[1], y + 2, str(q))
            c.drawRightString(col_x[2] + col_w[2], y + 2, f"{pu:.2f} DH")
            c.drawRightString(col_x[3] + col_w[3], y + 2, f"{st_val:.2f} DH")
            c.setFillColor(black)
            y -= 20

    # ── SEPARATOR ────────────────────────────────────────────────
    c.setStrokeColor(colors.HexColor("#cccccc"))
    c.setLineWidth(0.5)
    c.line(50, y, 530, y)
    c.setLineWidth(1)
    c.setStrokeColor(black)
    y -= 10

    # ── TOTALS BLOCK ─────────────────────────────────────────────
    tot_x, tot_w = 310, 220
    # Total row
    draw_rect_filled(tot_x, y - 4, tot_w, 22, colors.HexColor("#f0f4ff"))
    c.setFillColor(colors.HexColor("#333333"))
    c.setFont("Helvetica-Bold", 10)
    c.drawString(tot_x + 8, y + 4, "Montant Total:")
    c.drawRightString(tot_x + tot_w - 8, y + 4, f"{mt:.2f} DH")
    y -= 22
    # Payé row
    draw_rect_filled(tot_x, y - 4, tot_w, 22, light_green)
    c.setFillColor(green)
    c.drawString(tot_x + 8, y + 4, "Montant Payé:")
    c.drawRightString(tot_x + tot_w - 8, y + 4, f"{mr:.2f} DH")
    y -= 22
    # Restant / Réglé row
    if rest > 0:
        draw_rect_filled(tot_x, y - 6, tot_w, 26, light_red)
        c.setFillColor(red)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(tot_x + 8, y + 6, "Reste à Payer:")
        c.drawRightString(tot_x + tot_w - 8, y + 6, f"{rest:.2f} DH")
    else:
        draw_rect_filled(tot_x, y - 6, tot_w, 26, light_green)
        c.setFillColor(green)
        c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(tot_x + tot_w / 2, y + 6, "✓ RÉGLÉ INTÉGRALEMENT")
    y -= 26 + 12
    c.setFillColor(black)

    # ── STATUT BADGE ─────────────────────────────────────────────
    statut_map = {
        "Payee":               (green,  light_green,  "PAYÉE"),
        "Impayee":             (red,    light_red,    "IMPAYÉE"),
        "Partiellement payee": (orange, light_orange, "PARTIELLEMENT PAYÉE"),
    }
    st_color, st_bg, st_label = statut_map.get(
        facture.get("statut", ""), (colors.HexColor("#555555"), colors.HexColor("#eeeeee"), facture.get("statut", ""))
    )
    draw_rect_filled(50, y - 6, 200, 26, st_bg)
    c.setFillColor(st_color)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(58, y + 6, f"STATUT: {st_label}")
    y -= 26 + 14
    c.setFillColor(black)

    # ── PAYMENT HISTORY TABLE ─────────────────────────────────────
    if paiements:
        c.setFillColor(blue)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, y, "Historique des paiements")
        c.setStrokeColor(blue)
        c.setLineWidth(1)
        c.line(50, y - 3, 530, y - 3)
        c.setLineWidth(1)
        c.setStrokeColor(black)
        y -= 22

        # Mini table header
        draw_rect_filled(50, y - 4, 480, 20, blue)
        c.setFillColor(white)
        c.setFont("Helvetica-Bold", 9)
        c.drawString(56, y + 2, "#")
        c.drawString(80, y + 2, "Date")
        c.drawRightString(270, y + 2, "Montant")
        c.drawString(278, y + 2, "Type")
        c.drawString(360, y + 2, "Notes")
        y -= 20
        c.setFillColor(black)

        for i, p in enumerate(paiements):
            if y < 80:
                draw_rect_filled(0, 0, w, 40, blue)
                c.setFillColor(white)
                c.setFont("Helvetica", 8)
                c.drawCentredString(w / 2, 15, "Merci pour votre confiance ! — Généré par DentiPro")
                c.showPage()
                y = h - 50
            row_bg = colors.HexColor("#f8f9fa") if i % 2 == 0 else white
            draw_rect_filled(50, y - 4, 480, 20, row_bg)
            c.setFillColor(colors.HexColor("#333333"))
            c.setFont("Helvetica", 9)
            c.drawString(56, y + 2, str(i + 1))
            c.drawString(80, y + 2, fmt_date(p.get("date_paiement")))
            c.drawRightString(270, y + 2, f"{float(p.get('montant') or 0):.2f} DH")
            c.drawString(278, y + 2, str(p.get("type_paiement") or "-")[:12])
            c.drawString(360, y + 2, str(p.get("notes") or "-")[:28])
            y -= 20
        c.setFillColor(black)
        y -= 10

    # ── NOTES ────────────────────────────────────────────────────
    if facture.get("notes"):
        c.setFont("Helvetica-Bold", 10)
        c.drawString(50, y, "Notes:")
        c.setFillColor(colors.HexColor("#555555"))
        c.setFont("Helvetica", 10)
        c.drawString(50, y - 14, str(facture["notes"])[:100])
        c.setFillColor(black)

    # ── FOOTER ───────────────────────────────────────────────────
    draw_rect_filled(0, 0, w, 40, blue)
    c.setFillColor(white)
    c.setFont("Helvetica", 8)
    c.drawCentredString(w / 2, 16, "Merci pour votre confiance ! — Généré par DentiPro")
    c.setFillColor(black)

    c.save()
    buf.seek(0)
    return buf

def register_factures_routes(app):
    @app.get("/api/factures/next-numero")
    def fac_next():
        r = query_all(
            "SELECT MAX(CAST(SUBSTRING_INDEX(numero_facture, '-', -1) AS UNSIGNED)) as last_num FROM facture WHERE numero_facture LIKE 'FAC-%%'"
        )
        n = int(r[0]["last_num"] or 0) + 1
        return jsonify({"numero_facture": "FAC-" + str(n).zfill(5)})

    @app.get("/api/factures/stats/summary")
    def fac_stats():
        dd, df = request.args.get("date_debut"), request.args.get("date_fin")
        where, params = "", []
        if dd and df:
            where, params = "WHERE date_facture BETWEEN %s AND %s", [dd, df]
        elif dd:
            where, params = "WHERE date_facture >= %s", [dd]
        elif df:
            where, params = "WHERE date_facture <= %s", [df]
        sql = f"""SELECT COUNT(*) as total_factures, SUM(montant_total) as montant_total,
        SUM(montant_regle) as montant_regle, SUM(montant_total - montant_regle) as montant_restant,
        SUM(CASE WHEN statut='Payee' THEN 1 ELSE 0 END) as factures_payees,
        SUM(CASE WHEN statut='Impayee' THEN 1 ELSE 0 END) as factures_impayees,
        SUM(CASE WHEN statut='Partiellement payee' THEN 1 ELSE 0 END) as factures_partielles
        FROM facture {where}"""
        return jsonify(sj(query_all(sql, tuple(params))[0]))

    @app.get("/api/factures")
    def fac_list():
        pid = request.args.get("patient_id")
        st = request.args.get("statut")
        dd = request.args.get("date_debut")
        df = request.args.get("date_fin")
        sql = """SELECT f.*, pat.nom as patient_nom, pat.prenom as patient_prenom,
        pat.telephone as patient_telephone FROM facture f
        LEFT JOIN patient pat ON f.id_patient = pat.id_patient WHERE 1=1"""
        params = []
        if pid:
            sql += " AND f.id_patient = %s"
            params.append(pid)
        if st:
            sql += " AND f.statut = %s"
            params.append(st)
        if dd:
            sql += " AND f.date_facture >= %s"
            params.append(dd)
        if df:
            sql += " AND f.date_facture <= %s"
            params.append(df)
        sql += " ORDER BY f.date_facture DESC, f.id_facture DESC"
        return jsonify(sjl(query_all(sql, tuple(params))))

    @app.get("/api/factures/<int:fid>/pdf")
    def fac_pdf(fid):
        sql = """SELECT f.*, pat.nom as patient_nom, pat.prenom as patient_prenom,
        pat.telephone as patient_telephone, pat.address as patient_address, pat.cnie as patient_cnie
        FROM facture f LEFT JOIN patient pat ON f.id_patient = pat.id_patient WHERE f.id_facture=%s"""
        r = query_all(sql, (fid,))
        if not r:
            return jsonify({"error": "Facture non trouvée"}), 404
        fac = r[0]
        det = query_all(
            """SELECT fd.*, s.nom_service FROM facture_detail fd
            LEFT JOIN service s ON fd.id_service = s.id_service WHERE fd.id_facture=%s""",
            (fid,),
        )
        pay = query_all(
            "SELECT * FROM paiement WHERE id_facture = %s ORDER BY date_paiement ASC", (fid,)
        )
        buf = _pdf_facture(fac, det, pay or [])
        num = fac.get("numero_facture") or str(fid)
        return send_file(
            buf,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"facture-{num}.pdf",
        )

    @app.get("/api/factures/<int:fid>")
    def fac_one(fid):
        sql = """SELECT f.*, pat.nom as patient_nom, pat.prenom as patient_prenom,
        pat.telephone as patient_telephone, pat.address as patient_address
        FROM facture f LEFT JOIN patient pat ON f.id_patient = pat.id_patient WHERE f.id_facture=%s"""
        r = query_all(sql, (fid,))
        if not r:
            return jsonify({"error": "Facture non trouvée"}), 404
        facture = sj(r[0])
        det = query_all(
            """SELECT fd.*, s.nom_service FROM facture_detail fd
            LEFT JOIN service s ON fd.id_service = s.id_service WHERE fd.id_facture=%s""",
            (fid,),
        )
        facture["details"] = sjl(det)
        pay = query_all(
            "SELECT * FROM paiement WHERE id_facture = %s ORDER BY date_paiement ASC", (fid,)
        )
        facture["paiements"] = sjl(pay or [])
        return jsonify(facture)

    @app.post("/api/factures")
    def fac_add():
        b = request.get_json(silent=True) or {}
        st = b.get("statut") or "Impayee"
        iid = execute(
            """INSERT INTO facture
            (id_patient, id_user, numero_facture, date_facture, montant_total, montant_regle, statut, notes, motif, dent)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (
                b.get("id_patient"),
                b.get("id_user"),
                b.get("numero_facture"),
                b.get("date_facture"),
                b.get("montant_total"),
                b.get("montant_regle") or 0,
                st,
                b.get("notes") or None,
                b.get("motif") or None,
                b.get("dent") or None,
            ),
            fetch_last=True,
        )
        details = b.get("details") or []
        if details:
            rows = [
                (
                    iid,
                    d.get("id_service") or None,
                    d.get("description"),
                    d.get("quantite"),
                    d.get("prix_unitaire"),
                    d.get("sous_total"),
                )
                for d in details
            ]
            executemany(
                """INSERT INTO facture_detail (id_facture, id_service, description, quantite, prix_unitaire, sous_total)
                VALUES (%s,%s,%s,%s,%s,%s)""",
                rows,
            )
        excel_sync.fire_excel(excel_sync.append_facture, iid)
        return jsonify({"success": True, "id": iid})

    @app.put("/api/factures/<int:fid>")
    def fac_put(fid):
        b = request.get_json(silent=True) or {}
        details = b.get("details") or []
        final_total = b.get("montant_total")
        if details:
            final_total = sum(
                float(d.get("quantite") or 0) * float(d.get("prix_unitaire") or 0) for d in details
            )
        execute(
            """UPDATE facture SET id_patient=%s, id_user=%s, numero_facture=%s, date_facture=%s,
            montant_total=%s, montant_regle=%s, statut=%s, notes=%s, motif=%s, dent=%s WHERE id_facture=%s""",
            (
                b.get("id_patient"),
                b.get("id_user"),
                b.get("numero_facture"),
                b.get("date_facture"),
                final_total,
                b.get("montant_regle"),
                b.get("statut"),
                b.get("notes") or None,
                b.get("motif") or None,
                b.get("dent") or None,
                fid,
            ),
        )
        if details:
            execute("DELETE FROM facture_detail WHERE id_facture=%s", (fid,))
            rows = [
                (
                    fid,
                    d.get("id_service") or None,
                    d.get("description"),
                    d.get("quantite"),
                    d.get("prix_unitaire"),
                    float(d.get("quantite") or 0) * float(d.get("prix_unitaire") or 0),
                )
                for d in details
            ]
            executemany(
                """INSERT INTO facture_detail (id_facture, id_service, description, quantite, prix_unitaire, sous_total)
                VALUES (%s,%s,%s,%s,%s,%s)""",
                rows,
            )
            return jsonify({"success": True, "id": fid, "montant_total": final_total})
        return jsonify({"success": True, "id": fid})

    @app.delete("/api/factures/<int:fid>")
    def fac_del(fid):
        execute("DELETE FROM facture_detail WHERE id_facture=%s", (fid,))
        execute("DELETE FROM facture WHERE id_facture=%s", (fid,))
        return jsonify({"success": True})


def register_salle_routes(app):
    @app.get("/api/salle-attente/actuel")
    def sa_act():
        t = _today_local()
        r = query_all(
            """SELECT sa.*, p.nom, p.prenom, p.telephone FROM salle_attente sa
            JOIN patient p ON sa.id_patient = p.id_patient
            WHERE sa.date_attente = %s AND sa.statut = 'En cours' ORDER BY sa.position ASC LIMIT 1""",
            (t,),
        )
        if not r:
            return jsonify(None)
        return jsonify(sj(r[0]))

    @app.get("/api/salle-attente/file")
    def sa_file():
        t = _today_local()
        r = query_all(
            """SELECT sa.*, p.nom, p.prenom, p.telephone FROM salle_attente sa
            JOIN patient p ON sa.id_patient = p.id_patient
            WHERE sa.date_attente = %s AND sa.statut = 'En attente' ORDER BY sa.position ASC""",
            (t,),
        )
        return jsonify(sjl(r))

    @app.get("/api/salle-attente/stats")
    def sa_stats():
        t = _today_local()
        r = query_all(
            """SELECT COUNT(*) as total,
            SUM(CASE WHEN statut = 'En attente' THEN 1 ELSE 0 END) as en_attente,
            SUM(CASE WHEN statut = 'En cours' THEN 1 ELSE 0 END) as en_cours,
            SUM(CASE WHEN statut = 'Terminé' THEN 1 ELSE 0 END) as termine
            FROM salle_attente WHERE date_attente = %s""",
            (t,),
        )
        return jsonify(sj(r[0]))

    @app.post("/api/salle-attente/appeler-suivant")
    def sa_next():
        t = _today_local()
        cur = query_all(
            "SELECT id FROM salle_attente WHERE date_attente = %s AND statut = 'En cours'", (t,)
        )
        if cur:
            execute("UPDATE salle_attente SET statut = 'En attente' WHERE id = %s", (cur[0]["id"],))
        nxt = query_all(
            """SELECT sa.*, p.nom, p.prenom FROM salle_attente sa
            JOIN patient p ON sa.id_patient = p.id_patient
            WHERE sa.date_attente = %s AND sa.statut = 'En attente' ORDER BY sa.position ASC LIMIT 1""",
            (t,),
        )
        if not nxt:
            return jsonify({"success": False, "message": "Aucun patient en attente", "patient": None})
        p = nxt[0]
        execute("UPDATE salle_attente SET statut = 'En cours' WHERE id = %s", (p["id"],))
        return jsonify(
            {
                "success": True,
                "patient": sj(p),
                "message": f"Appeler: {p['nom']} {p['prenom']}",
            }
        )

    @app.post("/api/salle-attente/reset")
    def sa_reset():
        t = _today_local()
        execute("DELETE FROM salle_attente WHERE date_attente = %s", (t,))
        return jsonify({"success": True, "message": "Salle d'attente réinitialisée"})

    @app.post("/api/salle-attente/<int:sid>/terminer")
    def sa_term(sid):
        t = _today_local()
        chk = query_all(
            "SELECT * FROM salle_attente WHERE id = %s AND statut = 'En cours'", (sid,)
        )
        if not chk:
            return jsonify({"error": "Aucun patient en cours à terminer"}), 400
        patient = chk[0]
        execute("UPDATE salle_attente SET statut = 'Terminé' WHERE id = %s", (sid,))
        execute(
            """UPDATE salle_attente SET position = position - 1
            WHERE date_attente = %s AND position > %s AND statut = 'En attente'""",
            (t, patient["position"]),
        )
        execute(
            """UPDATE salle_attente SET position = position - 1
            WHERE date_attente = %s AND statut = 'En attente' AND position > %s""",
            (t, patient["position"]),
        )
        return jsonify({"success": True, "message": "Consultation terminée"})

    @app.get("/api/salle-attente")
    def sa_list():
        t = _today_local()
        r = query_all(
            """SELECT sa.*, p.nom, p.prenom, p.telephone FROM salle_attente sa
            JOIN patient p ON sa.id_patient = p.id_patient WHERE sa.date_attente = %s ORDER BY sa.position ASC""",
            (t,),
        )
        return jsonify(sjl(r))

    @app.post("/api/salle-attente")
    def sa_add():
        b = request.get_json(silent=True) or {}
        date_a = _today_local()
        lp = query_all(
            "SELECT MAX(position) as last_position FROM salle_attente WHERE date_attente = %s", (date_a,)
        )
        new_p = int(lp[0]["last_position"] or 0) + 1
        iid = execute(
            """INSERT INTO salle_attente (id_patient, id_rdv, date_attente, heure_arrivee, position, statut, notes)
            VALUES (%s,%s,%s,%s,%s,'En attente',%s)""",
            (
                b.get("id_patient"),
                b.get("id_rdv") or None,
                date_a,
                b.get("heure_arrivee"),
                new_p,
                b.get("notes") or None,
            ),
            fetch_last=True,
        )
        return jsonify(
            {
                "success": True,
                "id": iid,
                "position": new_p,
                "message": "Patient ajouté à la salle d'attente",
            }
        )

    @app.get("/api/salle-attente/<int:sid>")
    def sa_one(sid):
        r = query_all(
            """SELECT sa.*, p.nom, p.prenom, p.telephone FROM salle_attente sa
            JOIN patient p ON sa.id_patient = p.id_patient WHERE sa.id = %s""",
            (sid,),
        )
        if not r:
            return jsonify({"error": "Patient non trouvé dans la salle d'attente"}), 404
        return jsonify(sj(r[0]))

    @app.put("/api/salle-attente/<int:sid>")
    def sa_put(sid):
        b = request.get_json(silent=True) or {}
        execute(
            """UPDATE salle_attente SET id_patient=%s, id_rdv=%s, heure_arrivee=%s, notes=%s WHERE id=%s""",
            (
                b.get("id_patient"),
                b.get("id_rdv") or None,
                b.get("heure_arrivee"),
                b.get("notes") or None,
                sid,
            ),
        )
        return jsonify({"success": True, "message": "Patient modifié avec succès"})

    @app.patch("/api/salle-attente/<int:sid>/statut")
    def sa_statut(sid):
        b = request.get_json(silent=True) or {}
        st = b.get("statut")
        if st not in ("En attente", "En cours", "Terminé"):
            return (
                jsonify(
                    {"error": "Statut invalide. Utilisez: En attente, En cours, Terminé"}
                ),
                400,
            )
        execute("UPDATE salle_attente SET statut = %s WHERE id = %s", (st, sid))
        return jsonify({"success": True, "message": "Statut mis à jour"})

    @app.patch("/api/salle-attente/<int:sid>/position")
    def sa_pos(sid):
        b = request.get_json(silent=True) or {}
        pos = b.get("position")
        if not pos or int(pos) < 1:
            return jsonify({"error": "Position invalide"}), 400
        pos = int(pos)
        today = _today_local()
        cur = query_all("SELECT position FROM salle_attente WHERE id = %s", (sid,))
        if not cur:
            return jsonify({"error": "Patient non trouvé"}), 404
        old = int(cur[0]["position"])
        if old == pos:
            return jsonify({"success": True, "message": "Position inchangée"})
        if pos > old:
            execute(
                """UPDATE salle_attente SET position = position - 1
                WHERE date_attente = %s AND position > %s AND position <= %s""",
                (today, pos, old),
            )
        else:
            execute(
                """UPDATE salle_attente SET position = position + 1
                WHERE date_attente = %s AND position >= %s AND position < %s""",
                (today, pos, old),
            )
        execute("UPDATE salle_attente SET position = %s WHERE id = %s", (pos, sid))
        return jsonify({"success": True, "message": "Position mise à jour"})

    @app.delete("/api/salle-attente/<int:sid>")
    def sa_del(sid):
        r = query_all("SELECT position, date_attente FROM salle_attente WHERE id = %s", (sid,))
        if not r:
            return jsonify({"error": "Patient non trouvé"}), 404
        dp, da = r[0]["position"], r[0]["date_attente"]
        execute("DELETE FROM salle_attente WHERE id = %s", (sid,))
        execute(
            "UPDATE salle_attente SET position = position - 1 WHERE date_attente = %s AND position > %s",
            (da, dp),
        )
        return jsonify({"success": True, "message": "Patient supprimé de la salle d'attente"})


def register_schema_routes(app):
    @app.get("/api/schema-dentaire/<int:patient_id>")
    def sch_get(patient_id):
        r = query_all(
            "SELECT * FROM schema_dentaire WHERE id_patient = %s ORDER BY numero_dent ASC",
            (patient_id,),
        )
        return jsonify({"success": True, "data": sjl(r)})

    @app.post("/api/schema-dentaire")
    def sch_post():
        b = request.get_json(silent=True) or {}
        if not b.get("id_patient") or b.get("numero_dent") is None:
            return jsonify({"success": False, "message": "id_patient et numero_dent sont requis"}), 400
        execute(
            """INSERT INTO schema_dentaire (id_patient, numero_dent, etat, notes)
            VALUES (%s,%s,%s,%s) ON DUPLICATE KEY UPDATE etat=VALUES(etat), notes=VALUES(notes)""",
            (
                b["id_patient"],
                b["numero_dent"],
                b.get("etat") or "saine",
                b.get("notes") or None,
            ),
        )
        return jsonify({"success": True, "message": "Dent enregistrée"})

    @app.delete("/api/schema-dentaire/<int:patient_id>/<int:numero_dent>")
    def sch_del(patient_id, numero_dent):
        execute(
            "DELETE FROM schema_dentaire WHERE id_patient = %s AND numero_dent = %s",
            (patient_id, numero_dent),
        )
        return jsonify({"success": True})

    @app.get("/api/rdv-dents/<int:id_rdv>")
    def rdv_dents_get(id_rdv):
        r = query_all(
            """SELECT rd.numero_dent, sd.etat, sd.notes FROM rdv_dents rd
            LEFT JOIN schema_dentaire sd ON rd.id_patient = sd.id_patient AND rd.numero_dent = sd.numero_dent
            WHERE rd.id_rdv = %s ORDER BY rd.numero_dent ASC""",
            (id_rdv,),
        )
        return jsonify({"success": True, "data": sjl(r)})

    @app.post("/api/rdv-dents")
    def rdv_dents_post():
        b = request.get_json(silent=True) or {}
        if not b.get("id_rdv") or not b.get("id_patient") or not isinstance(b.get("dents"), list):
            return jsonify({"success": False, "message": "Données manquantes"}), 400
        execute("DELETE FROM rdv_dents WHERE id_rdv = %s", (b["id_rdv"],))
        dents = b["dents"]
        if dents:
            rows = [(b["id_rdv"], b["id_patient"], d) for d in dents]
            executemany(
                "INSERT INTO rdv_dents (id_rdv, id_patient, numero_dent) VALUES (%s,%s,%s)", rows
            )
            return jsonify(
                {"success": True, "message": f"{len(dents)} dent(s) associée(s)"}
            )
        return jsonify({"success": True, "message": "Dents mises à jour"})


def _linked_partner(uid: int):
    rows = query_all(
        "SELECT linked_user_id FROM utilisateur WHERE id_user=%s", (uid,)
    )
    if not rows:
        return None
    lid = rows[0].get("linked_user_id")
    if lid is None:
        return None
    try:
        return int(lid)
    except (TypeError, ValueError):
        return None


def register_chat_routes(app):
    @app.get("/api/chat/unread/<int:id_user>")
    def chat_unread(id_user):
        partner = _linked_partner(id_user)
        if partner:
            r = query_all(
                "SELECT COUNT(*) as count FROM chat_message WHERE id_user=%s AND lu=0",
                (partner,),
            )
        else:
            r = query_all(
                "SELECT COUNT(*) as count FROM chat_message WHERE id_user != %s AND lu = 0",
                (id_user,),
            )
        return jsonify(sj(r[0]))

    @app.post("/api/chat/fichier")
    def chat_file():
        if "fichier" not in request.files:
            return jsonify({"error": "Fichier manquant"}), 400
        f = request.files["fichier"]
        ALLOWED_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.pdf', '.doc', '.docx'}
        ext = Path(f.filename or "").suffix
        if ext not in ALLOWED_EXTENSIONS:
            return jsonify({"error": "Type de fichier non autorisé."}), 400
        name = f"{int(datetime.now().timestamp() * 1000)}_{random.randint(0, 10**6)}{ext}"
        path = UPLOAD_CHAT / name
        f.save(path)
        b = request.form
        uid = b.get("id_user")
        url = f"/uploads/chat/{name}"
        iid = execute(
            """INSERT INTO chat_message (id_user, message, fichier_url, fichier_nom, fichier_type)
            VALUES (%s,%s,%s,%s,%s)""",
            (
                uid,
                b.get("message") or f.filename,
                url,
                f.filename,
                f.mimetype or "application/octet-stream",
            ),
            fetch_last=True,
        )
        return jsonify({"success": True, "id": iid, "url": url})

    @app.patch("/api/chat/lu/<int:id_user>")
    def chat_lu(id_user):
        partner = _linked_partner(id_user)
        if partner:
            execute(
                "UPDATE chat_message SET lu=1 WHERE id_user=%s AND lu=0", (partner,)
            )
        else:
            execute("UPDATE chat_message SET lu=1 WHERE id_user != %s", (id_user,))
        return jsonify({"success": True})

    @app.get("/api/chat")
    def chat_list():
        uid = request.args.get("id_user", type=int)
        if uid:
            partner = _linked_partner(uid)
            if partner:
                r = query_all(
                    """SELECT m.*, u.nom, u.prenom, u.role FROM chat_message m
                    JOIN utilisateur u ON m.id_user = u.id_user
                    WHERE m.id_user IN (%s,%s)
                    ORDER BY m.created_at DESC LIMIT 100""",
                    (uid, partner),
                )
            else:
                r = query_all(
                    """SELECT m.*, u.nom, u.prenom, u.role FROM chat_message m
                    JOIN utilisateur u ON m.id_user = u.id_user
                    WHERE LOWER(u.role) IN ('dentiste','secretaire')
                    ORDER BY m.created_at DESC LIMIT 100""",
                )
        else:
            r = query_all(
                """SELECT m.*, u.nom, u.prenom, u.role FROM chat_message m
                JOIN utilisateur u ON m.id_user = u.id_user
                ORDER BY m.created_at DESC LIMIT 50""",
            )
        return jsonify(list(reversed(sjl(r))))

    @app.post("/api/chat")
    def chat_post():
        b = request.get_json(silent=True) or {}
        iid = execute(
            "INSERT INTO chat_message (id_user, message) VALUES (%s,%s)",
            (b.get("id_user"), b.get("message")),
            fetch_last=True,
        )
        return jsonify({"success": True, "id": iid})


def register_export_routes(app):
    @app.get("/api/export/excel")
    def export_xlsx():
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        try:
            patients = query_all("SELECT * FROM patient ORDER BY nom ASC")
            rdvs = query_all(
                "SELECT rv.*, CONCAT(p.nom,' ',p.prenom) AS pn FROM rendez_vous rv LEFT JOIN patient p ON rv.id_patient=p.id_patient ORDER BY rv.date_rdv DESC"
            )
            factures = query_all(
                "SELECT f.*, CONCAT(p.nom,' ',p.prenom) AS pn FROM facture f LEFT JOIN patient p ON f.id_patient=p.id_patient ORDER BY f.date_facture DESC"
            )
            paiements = query_all(
                "SELECT py.*, CONCAT(p.nom,' ',p.prenom) AS pn, f.numero_facture FROM paiement py LEFT JOIN patient p ON py.id_patient=p.id_patient LEFT JOIN facture f ON py.id_facture=f.id_facture ORDER BY py.date_paiement DESC"
            )
            wb = Workbook()
            wb.remove(wb.active)
            thin = Side(style="thin", color="FFD1D5DB")

            def hdr(ws, titles, color):
                ws.append(titles)
                for i, _ in enumerate(titles, 1):
                    cell = ws.cell(1, i)
                    cell.fill = PatternFill("solid", fgColor=color)
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(thin, thin, thin, thin)

            ws_p = wb.create_sheet("Patients")
            hdr(
                ws_p,
                [
                    "ID",
                    "Nom",
                    "Prénom",
                    "Sexe",
                    "Tél",
                    "CIN",
                    "Naissance",
                    "Email",
                    "Créé",
                ],
                "FF1A73E8",
            )
            for p in patients:
                ws_p.append(
                    [
                        p.get("id_patient"),
                        p.get("nom"),
                        p.get("prenom"),
                        p.get("sexe"),
                        p.get("telephone"),
                        p.get("cnie"),
                        str(p.get("date_naissance") or ""),
                        p.get("email"),
                        str(p.get("created_at") or ""),
                    ]
                )
            ws_r = wb.create_sheet("Rendez-vous")
            hdr(
                ws_r,
                ["ID", "Patient", "Date", "Heure", "Motif", "Dent", "Statut"],
                "FF0F9D58",
            )
            for r in rdvs:
                ws_r.append(
                    [
                        r.get("id_rdv"),
                        r.get("pn"),
                        str(r.get("date_rdv") or ""),
                        str(r.get("heure_rdv") or ""),
                        r.get("motif"),
                        r.get("dent"),
                        r.get("statut"),
                    ]
                )
            ws_f = wb.create_sheet("Factures")
            hdr(
                ws_f,
                ["ID", "Patient", "N°", "Date", "Total", "Payé", "Statut"],
                "FFE65100",
            )
            for frow in factures:
                tot = float(frow.get("montant_total") or 0)
                pay = float(frow.get("montant_regle") or 0)
                ws_f.append(
                    [
                        frow.get("id_facture"),
                        frow.get("pn"),
                        frow.get("numero_facture"),
                        str(frow.get("date_facture") or ""),
                        tot,
                        pay,
                        frow.get("statut"),
                    ]
                )
            ws_pay = wb.create_sheet("Paiements")
            hdr(
                ws_pay,
                ["ID", "Patient", "Montant", "Type", "Date", "Facture"],
                "FF7B1FA2",
            )
            for p in paiements:
                ws_pay.append(
                    [
                        p.get("id_paiement"),
                        p.get("pn"),
                        float(p.get("montant") or 0),
                        p.get("type_paiement"),
                        str(p.get("date_paiement") or ""),
                        p.get("numero_facture"),
                    ]
                )
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            fn = f"DentiPro_Export_{date.today().isoformat()}.xlsx"
            return send_file(
                bio,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=fn,
            )
        except Exception as e:
            return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    application = create_app()
    print("DentiPro (Flask) sur http://localhost:3000")
    application.run(host="0.0.0.0", port=3000, threaded=True)